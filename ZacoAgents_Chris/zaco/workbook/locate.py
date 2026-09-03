"""Finding the data sheet and its columns by header text, never by position.

The brief describes 21 columns, A to U. The operator's actual book has **23**: `Buyer note` is
inserted at C and `Packhouse` at V, so every column letter in the brief is wrong for the real
file. The data is also on `Sheet1`, the *third* sheet, after `Cover` and `Rates`.

A system that indexes by letter appends a book that opens without complaint and has the price
under the description and the notes under the status. So nothing here counts columns: the
header row is read, and every column is found by the words in it. A workbook whose columns
have been reordered, or which has grown another one, keeps working. One with no recognisable
header row is refused with an explanation rather than guessed at.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from dataclasses import field as dc_field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#: Every column the system reads or writes, and the header text that identifies it. Matching is
#: case- and space-insensitive, because a header retyped by hand rarely comes back identical.
COLUMNS: dict[str, str] = {
    "dn": "DN",
    "market_agent": "Market Agent",
    "completed": "Completed",
    "date": "Date",
    "stm_no": "STM No",
    "description": "Description",
    "qty_received": "Qty Received",
    "opening_stock": "Opening Stock",
    "frui_price_per_crt": "Frui Price/crt",
    "cartons_sold": "Cartons Sold",
    "baby_stock": "Baby Stock",
    "price": "Price",
    "gross_total": "Gross Total",
    "nett_total": "Nett Total",
    "nett_price_per_crt": "Nett Price/crt",
    "z_rand_per_crt": "Z R/crt",
    "z_total": "Z Total",
    "markup_percent": "% Markup",
    "frui_curr_sales_value": "Frui Curr Sales Value",
    "status": "Status",
    "notes": "NOTES",
}

#: Without these there is no data sheet worth appending to, and a guess would be worse than a
#: refusal. Everything else in COLUMNS may legitimately be missing from an older book.
REQUIRED = ("dn", "date", "stm_no", "description", "qty_received", "cartons_sold")

MAX_HEADER_SEARCH_ROWS = 10


class WorkbookShapeError(Exception):
    """The file has no sheet that looks like the account sales book (section 5)."""


def _normalise(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


@dataclass(frozen=True)
class SheetLayout:
    """Where everything is in one particular book."""

    sheet_name: str
    header_row: int
    first_data_row: int
    last_data_row: int
    columns: dict[str, int]
    """Field name to 1-based column index, for the columns this book actually has."""

    unknown_headers: dict[str, int]
    """Columns the book has that the system does not write. `Buyer note` and `Packhouse` land
    here, and they are the reason nothing may be written by position."""

    headers: dict[str, str] = dc_field(default_factory=dict)
    """Field name to the header text **as this book writes it**.

    Kept rather than re-derived from `COLUMNS`, because the point of showing an operator where
    the columns went is to show them their own words. A book that says `STM NO` should say
    `STM NO` back, not the `STM No` this module happens to match against.
    """

    def index(self, field: str) -> int | None:
        return self.columns.get(field)

    def letter(self, field: str) -> str | None:
        column = self.columns.get(field)
        return None if column is None else get_column_letter(column)

    @property
    def row_count(self) -> int:
        return max(0, self.last_data_row - self.first_data_row + 1)


def _score_row(cells: list[Any]) -> tuple[dict[str, int], dict[str, int], dict[str, str]]:
    wanted = {_normalise(text): field for field, text in COLUMNS.items()}
    found: dict[str, int] = {}
    unknown: dict[str, int] = {}
    headers: dict[str, str] = {}
    for offset, value in enumerate(cells, start=1):
        text = _normalise(value)
        if not text:
            continue
        field = wanted.get(text)
        if field is not None and field not in found:
            found[field] = offset
            headers[field] = str(value).strip()
        elif field is None:
            unknown[str(value).strip()] = offset
    return found, unknown, headers


def locate(path: Path) -> SheetLayout:
    """Find the data sheet in a workbook, or refuse and say why.

    Every sheet is examined, not just the first, and the best-matching header row wins. That is
    what makes `Cover` and `Rates` sitting in front of the data a non-event rather than a bug.
    """
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        best: SheetLayout | None = None
        best_score = 0
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SEARCH_ROWS, values_only=True),
                start=1,
            ):
                found, unknown, headers = _score_row(list(row))
                score = len(found)
                if score > best_score and all(field in found for field in REQUIRED):
                    best_score = score
                    best = SheetLayout(
                        sheet_name=sheet.title,
                        header_row=row_number,
                        first_data_row=row_number + 1,
                        last_data_row=max(sheet.max_row, row_number),
                        columns=found,
                        unknown_headers=unknown,
                        headers=headers,
                    )
        if best is None:
            raise WorkbookShapeError(
                f"{path.name} has no sheet with a recognisable header row. The system looks for "
                f"a row naming at least {', '.join(COLUMNS[f] for f in REQUIRED)}, in any order "
                "and on any sheet. Nothing was written."
            )
        return best
    finally:
        workbook.close()


@dataclass(frozen=True)
class BookRow:
    """One existing row of the operator's book, read leniently."""

    row_number: int
    dn: str | None
    stm_no: str | None
    description: str | None
    date: object | None

    cells: dict[str, str] = dc_field(default_factory=dict)
    """Every column of the row as it reads on screen, keyed by **column letter**.

    By letter and not by field name because the operator's own columns -- `Buyer note`,
    `Packhouse`, `NOTES` -- have no field name at all, and they are as much a part of the row as
    the ones this system knows. Only filled when `read_rows` is asked for it.
    """

    formulas: dict[str, str] = dc_field(default_factory=dict)
    """The formula behind a cell, where there is one, keyed by column letter.

    A spreadsheet cell holds a formula *and* the result Excel cached the last time it saved.
    openpyxl does not calculate, so a row this system appended and Excel has not yet opened has
    a formula and no cached result. Keeping both means such a row can show its formula rather
    than an empty cell, and can be told apart from one that is genuinely blank.
    """


def _text(value: object) -> str | None:
    """Read a cell as text without caring what type it was stored as.

    Column A is a number in the live book and column E is sometimes a number and sometimes a
    reference like `5644200/1`, so anything that insists on a type here starts throwing on a
    real file. A float that is a whole number loses its `.0`, because `14690.0` matches nothing.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _show(value: object) -> str:
    """A cell as the operator would read it, not as Python happens to hold it."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time.min else value.isoformat(" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def read_rows(
    path: Path, layout: SheetLayout | None = None, *, with_cells: bool = False
) -> list[BookRow]:
    """The existing rows, for the DN join and the series. Formulas are not evaluated.

    `with_cells` additionally reads every column of every row, for drawing the book on screen.
    It costs a second pass over the file -- once for what Excel last calculated and once for the
    formulas behind it -- so the callers that only want delivery notes do not pay for it.
    """
    layout = layout or locate(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    written: dict[int, list[Any]] = {}
    if with_cells:
        formulas = load_workbook(path, data_only=False, read_only=True)
        try:
            written = {
                number: list(row)
                for number, row in enumerate(
                    formulas[layout.sheet_name].iter_rows(
                        min_row=layout.first_data_row, values_only=True
                    ),
                    start=layout.first_data_row,
                )
            }
        finally:
            formulas.close()
    try:
        sheet = workbook[layout.sheet_name]
        rows: list[BookRow] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=layout.first_data_row, values_only=True),
            start=layout.first_data_row,
        ):
            cells = list(row)

            def cell(field: str, cells: list[Any] = cells) -> object | None:
                index = layout.index(field) if layout else None
                return cells[index - 1] if index is not None and index <= len(cells) else None

            if all(value is None for value in cells):
                continue
            rows.append(
                BookRow(
                    row_number=row_number,
                    dn=_text(cell("dn")),
                    stm_no=_text(cell("stm_no")),
                    description=_text(cell("description")),
                    date=cell("date"),
                    cells=_by_letter(cells, written.get(row_number, [])),
                    formulas=_formulas_by_letter(written.get(row_number, [])),
                )
            )
        return rows
    finally:
        workbook.close()


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _by_letter(calculated: list[Any], written: list[Any]) -> dict[str, str]:
    """What each cell shows: the value Excel cached, or the formula when it cached none.

    A row this system appended and nobody has opened in Excel yet has formulas and no cached
    results at all. Showing its formula is truer than showing it empty, and it is the same thing
    the append preview shows for the same cells.
    """
    if not written:
        return {}
    out: dict[str, str] = {}
    for offset in range(max(len(calculated), len(written))):
        value = calculated[offset] if offset < len(calculated) else None
        if value is None and offset < len(written):
            value = written[offset]
        text = _show(value)
        if text:
            out[get_column_letter(offset + 1)] = text
    return out


def _formulas_by_letter(written: list[Any]) -> dict[str, str]:
    return {
        get_column_letter(offset + 1): value
        for offset, value in enumerate(written)
        if _is_formula(value)
    }
