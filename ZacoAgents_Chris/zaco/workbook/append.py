"""Writing resolved rows into the operator's live book (section 5).

Three rules shape everything here.

**Nothing is located by position.** The brief describes 21 columns, A to U. The operator's real
book has 23: `Buyer note` at C and `Packhouse` at V push every letter after B along, so `Baby
Stock` is `=I{r}-K{r}` in the live file and not the `=H{r}-J{r}` the brief prints. Formulas are
therefore built from letters `locate()` resolved, never from letters written down here.

**A computed value is never written into a formula column.** Eight columns belong to the
operator. The system writes the formula for the new row number and lets Excel do the arithmetic,
so an operator who corrects one input sees everything downstream follow.

**Rows are appended, never rebuilt.** The `NOTES` column -- the operator's own margin -- is not
read, not derived and not written.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from zaco.domain.model import StagedRound, display_account_sale
from zaco.money.allocate import CannotAllocateError, allocate
from zaco.resolve.reconcile import reconcile
from zaco.workbook import snapshot
from zaco.workbook.locate import SheetLayout, WorkbookShapeError, locate

#: The eight columns that belong to the operator, written as formulas for the new row number.
#: The placeholders are *field names*; the letters are filled in from the book being written.
FORMULAS: dict[str, str] = {
    "frui_price_per_crt": '=IFERROR({nett_price_per_crt}{r}*70%,"-")',
    "baby_stock": "={opening_stock}{r}-{cartons_sold}{r}",
    "gross_total": "=SUM({cartons_sold}{r}*{price}{r})",
    "nett_price_per_crt": '=IFERROR({nett_total}{r}/{cartons_sold}{r},"-")',
    "z_rand_per_crt": '=IFERROR({nett_price_per_crt}{r}-{frui_price_per_crt}{r},"-")',
    "z_total": "={nett_total}{r}-{frui_curr_sales_value}{r}",
    "markup_percent": '=IFERROR({z_rand_per_crt}{r}/{nett_price_per_crt}{r},"-")',
    "frui_curr_sales_value": "={frui_price_per_crt}{r}*{cartons_sold}{r}",
}

#: The operator's own margin. Read back and written back unchanged means: left alone entirely.
NEVER_WRITTEN = frozenset({"notes"})

#: What the price column is rounded to before it is written.
#:
#: Not a stylistic choice. `openpyxl` converts a `Decimal` to a float on the way into the file,
#: so `400 / 3` -- an exact 28-digit Decimal in this system -- is stored as
#: `133.33333333333331` and the tail is gone whatever we do. Excel keeps 15 significant digits
#: of its own, so there is no precision here to protect: rounding to the five decimals this
#: column already displays writes a number that survives the round trip exactly, instead of one
#: that changes quietly between the system and the file.
#:
#: Money itself is untouched by this. Every other currency column is two decimals and exact.
PRICE_PLACES = Decimal("0.00001")

#: Below half a cent the money columns round identically, so the difference is not worth saying.
WORTH_SAYING = Decimal("0.005")

INCOMPLETE = "Incomplete"


class AppendRefusedError(Exception):
    """The append was not attempted, and the file was not opened."""


@dataclass
class PlannedRow:
    """One row as it will be written, before any of it touches the file."""

    delivery_id: str | None
    consignment_id: str | None
    product: str
    account_sale: str
    values: dict[str, object] = field(default_factory=dict)
    blanks: dict[str, str] = field(default_factory=dict)
    """Field name to a short label saying why that cell is empty.

    Per cell rather than per row, because the grid puts the reason in the cell it belongs to.
    One sentence attached to the whole row leaves the reader matching it back to a column.
    """

    notes: list[str] = field(default_factory=list)
    """The same reasons at length, for the list beneath the grid."""

    blocked: list[str] = field(default_factory=list)
    """What is missing before this row could be written at all."""

    counted_with: int | None = None
    """Offset of the row that carries this consignment's Qty Received, when it is not this one.

    An offset rather than a row number, because the row number depends on where in the book the
    append lands and that is not known until the file is open. The preview turns it into the
    number the operator will actually see.
    """

    @property
    def is_writable(self) -> bool:
        return not self.blocked

    def blank(self, name: str, label: str, detail: str = "") -> None:
        self.values[name] = None
        self.blanks[name] = label
        if detail:
            self.notes.append(detail)


@dataclass
class AppendPlan:
    rows: list[PlannedRow] = field(default_factory=list)
    refusals: list[str] = field(default_factory=list)
    """Reasons the whole append cannot proceed. Non-empty means nothing is written."""

    @property
    def is_writable(self) -> bool:
        return bool(self.rows) and not self.refusals and all(r.is_writable for r in self.rows)


@dataclass
class AppendResult:
    first_row: int
    last_row: int
    saved_as: snapshot.Snapshot
    sheet_name: str
    letters: dict[str, str]


def _number_or_text(value: str | None) -> object | None:
    """D7: the bare number where one exists, the full reference where it does not.

    `382405` goes in as a number so it sorts and filters beside the operator's existing rows;
    `5644200/1` goes in as text, because dropping the suffix would merge two separate April
    payment runs worth R5,100 and R3,230 into one.
    """
    if value is None:
        return None
    return int(value) if value.isdigit() else value


def plan(resolved: Any) -> AppendPlan:
    """Work out every cell, and every reason not to write one, without opening the workbook.

    A row that cannot be written is still built and still shown, with the missing cells labelled.
    Dropping it and listing the reason separately gives an operator a short grid and a long
    complaint, and leaves them working out which row the complaint belongs to.
    """
    from zaco.db.models import RoundStatus

    built = AppendPlan()
    if resolved.round.status not in (RoundStatus.RESOLVED.value, RoundStatus.APPENDED.value):
        built.refusals.append(
            f"Round {resolved.round.id} is {resolved.round.status}. Only a round whose queue has "
            "been closed can be appended."
        )
    if resolved.round.appended_at is not None:
        built.refusals.append(
            f"Round {resolved.round.id} was already appended on "
            f"{resolved.round.appended_at:%d %b %Y at %H:%M}. Rows are appended, never rebuilt, "
            "so appending it again would write every row a second time."
        )
    if not resolved.is_clear:
        built.refusals.append(resolved.blocking_reason or "The queue is still open.")

    staged = resolved.staged
    counted: set[str] = set()
    first_row_of: dict[str, int] = {}
    rows_per_sale: dict[str, int] = {}
    for row in staged.rows:
        rows_per_sale[row.account_sale] = rows_per_sale.get(row.account_sale, 0) + 1
    nett_shares, nett_refusals = row_netts(staged)

    for offset, row in enumerate(staged.rows):
        note = resolved.approved.get(row.delivery_id or "")
        position = resolved.ledger.for_row(row)
        sale = staged.account_sales.get(row.account_sale)
        planned = PlannedRow(
            delivery_id=row.delivery_id,
            consignment_id=row.consignment_id,
            product=row.product.display_name,
            account_sale=row.account_sale,
        )

        # --- column A, and the date that hangs off it -----------------------------------------
        if note is None:
            planned.blocked.append(
                f"{row.delivery_id or '(no delivery)'} has no approved delivery note"
            )
            planned.blank(
                "dn",
                "DN not captured",
                f"{row.delivery_id or 'This row'} has no approved delivery note, so column A and "
                "the date grouped on it are both waiting on the queue.",
            )
        elif note.dn:
            planned.values["dn"] = _number_or_text(note.dn)
        else:
            # D11: visibly empty with the reason attached is a recorded answer, and a different
            # thing from a cell nobody reached.
            planned.blank(
                "dn",
                "no DN, recorded",
                f"{row.delivery_id}: no delivery note, recorded on purpose -- "
                f"{note.operator_reason or note.reasoning}",
            )

        planned.values["market_agent"] = row.agent
        planned.values["completed"] = INCOMPLETE

        # The date is the earliest across every row sharing a delivery note, so until the note
        # is approved this row's own earliest date is a guess that approving one could move.
        # Showing it anyway would put a figure in the preview that the append might not write.
        when = (resolved.grouping_dates.get(note.dn) if note and note.dn else None) or (
            row.earliest_date if note is not None else None
        )
        if when is not None:
            planned.values["date"] = when
        elif note is None:
            planned.blank(
                "date",
                "awaits DN",
                "The date is the earliest across every row sharing a delivery note, so approving "
                "one can move it earlier. It is not settled until the note is.",
            )
        else:
            planned.blank(
                "date",
                "no date in the documents",
                "No document in this round carries a date for this row.",
            )

        # Through `display_account_sale` even when no payment run accounts for the row: the
        # column is the operator's STM No either way, and `PRE*BT*390100` sitting between two
        # bare numbers is the sort of thing that stops a filter working.
        planned.values["stm_no"] = _number_or_text(display_account_sale(row.account_sale))

        if row.product.short_code is not None:
            planned.values["description"] = row.product.short_code
        else:
            planned.blocked.append(f"{row.product.display_name} has no short code")
            planned.blank(
                "description",
                "code unmapped",
                f"{row.product.display_name} has no short code. Section 7: the code is the "
                "operator's own and is not derivable from any report.",
            )

        # --- what was sent, counted once --------------------------------------------------------
        # It belongs to the consignment, not to the row. A consignment that sold under two
        # account sales makes two rows, and writing its cartons on both would double them in
        # every column that sums; a delivery holding two products makes two consignments, and
        # writing the delivery total on each would double them again the other way.
        consignment_id = row.consignment_id or ""
        consignment = next(
            (c for c in staged.consignments if c.consignment_id == consignment_id), None
        )
        if consignment_id and consignment_id not in counted:
            counted.add(consignment_id)
            first_row_of[consignment_id] = offset
            sent = consignment.qty_sent if consignment is not None else None
            if sent is not None:
                planned.values["qty_received"] = sent
            else:
                # Absent is not zero (section 6). The consignment report carries what was sent
                # and the daily sales file does not, so a round without one simply does not know.
                planned.blank(
                    "qty_received",
                    "not reported",
                    "No document in this round says what this consignment was sent. That is "
                    "absent, not nought.",
                )
        else:
            planned.counted_with = first_row_of.get(consignment_id)
            planned.blank(
                "qty_received",
                "counted above",
                "What was sent is sent once, however many account sales it sells under, so it "
                "stays on the first row of the consignment.",
            )

        if position is not None and position.opening is not None:
            planned.values["opening_stock"] = position.opening
        else:
            planned.blank(
                "opening_stock",
                "not known",
                "Nothing says what was on the floor when this row started selling, and a running "
                "balance cannot start from a guess.",
            )

        planned.values["cartons_sold"] = row.cartons.net
        price = _priced(row, planned)
        if price is not None:
            planned.values["price"] = price
        else:
            planned.blank(
                "price",
                "nothing net sold",
                "Price is the money over the net cartons, and this row nets nought.",
            )

        # --- the payment side ---------------------------------------------------------------------
        # A row is delivery x product x account sale, so it names exactly one account sale and
        # takes one share. Section 8's "a row that names more than one" cannot arise at this
        # grain; it would if the grain were the consignment.
        named = display_account_sale(row.account_sale)
        if sale is None or sale.nett is None:
            planned.blank(
                "nett_total",
                "no payment run",
                f"No payment document in this round accounts for {named}.",
            )
        elif (share := nett_shares.get(offset)) is not None:
            planned.values["nett_total"] = share
            if rows_per_sale[row.account_sale] > 1:
                planned.notes.append(
                    f"{sale.display_number} pays R{sale.nett:,.2f} across "
                    f"{rows_per_sale[row.account_sale]} rows. This row's share of it is "
                    f"R{share:,.2f}, worked out by sales value; the shares sum to the payment "
                    f"exactly. Apportioned, not printed -- type over it if the split is wrong."
                )
        else:
            planned.blank(
                "nett_total",
                f"split across {rows_per_sale[row.account_sale]} rows",
                nett_refusals.get(
                    row.account_sale,
                    f"{sale.display_number} pays R{sale.nett:,.2f} across "
                    f"{rows_per_sale[row.account_sale]} rows and the split could not be made.",
                ),
            )

        paid = sale.date_paid if sale else None
        if paid is not None:
            planned.values["status"] = f"{paid:%d.%m}"
        else:
            planned.blank("status", "not yet paid")

        built.refusals.extend(planned.blocked)
        built.rows.append(planned)

    if not built.rows and not built.refusals:
        built.refusals.append("This round produced no rows, so there is nothing to append.")
    return built


def row_netts(staged: StagedRound) -> tuple[dict[int, Decimal], dict[str, str]]:
    """What each row is paid, and per account sale the reason a row is not (section 8).

    "An account sale settles several rows at once, so its Nett is split between them by sales
    value... The shares must add up to the payment exactly."

    **A group of several rows is filled only when the two sides agree.** One that does not would
    otherwise receive a Nett for produce not all paid for yet. That rule also does the work of a
    guard nobody has to write: a group whose rows are split across two rounds can never be fully
    matched inside either of them, so neither round writes a figure and the payment is not
    settled twice.

    **A group of one row is filled either way**, and that is not an inconsistency. The rule
    exists to stop untrustworthy *proportions* being used, and a single row has no proportions --
    the whole Nett belongs to the only row under the account sale however the gross is disputed.
    Withholding it there would leave a cell empty over a disagreement about a different figure.

    The workbook writes these; settlement reads the same map, so a supplier is never settled
    against a figure the book does not hold.
    """
    shares: dict[int, Decimal] = {}
    refusals: dict[str, str] = {}

    grouped: dict[str, list[int]] = {}
    for offset, row in enumerate(staged.rows):
        grouped.setdefault(row.account_sale, []).append(offset)

    agreed = {r.account_sale: r for r in reconcile(staged)}
    for number, offsets in grouped.items():
        sale = staged.account_sales.get(number)
        if sale is None or sale.nett is None:
            continue
        if len(offsets) == 1:
            shares[offsets[0]] = sale.nett
            continue
        found = agreed.get(number)
        if found is None or not found.agrees:
            refusals[number] = (
                f"{sale.display_number} pays R{sale.nett:,.2f} across {len(offsets)} rows, and "
                f"the two sides do not agree: {found.note if found else 'nothing reconciles it'} "
                f"Only a fully matched group is filled, because a part-paid one would take a "
                f"Nett for produce nobody has been paid for yet."
            )
            continue
        try:
            portions = allocate(sale.nett, [staged.rows[i].value for i in offsets])
        except CannotAllocateError as refusal:
            refusals[number] = (
                f"{sale.display_number} pays R{sale.nett:,.2f} across {len(offsets)} rows and "
                f"the split cannot be made: {refusal}"
            )
            continue
        for offset, portion in zip(offsets, portions, strict=True):
            shares[offset] = portion
    return shares, refusals


def _priced(row: Any, planned: PlannedRow) -> Decimal | None:
    """The price, rounded to what the column carries, saying so when the money no longer lands.

    Section 5 asks for a price such that `Cartons Sold * Price` recovers the money. Where the
    value does not divide by the cartons -- R400 across 3 -- no price at any precision does that,
    and the row that says so is worth more than a figure that looks exact and is not.
    """
    price = row.price
    if price is None:
        return None
    rounded: Decimal = price.quantize(PRICE_PLACES, rounding=ROUND_HALF_UP)
    shortfall: Decimal = row.value - (rounded * row.cartons.net)
    if abs(shortfall) >= WORTH_SAYING:
        planned.notes.append(
            f"Cartons Sold x Price recovers R{rounded * row.cartons.net:,.2f} against the "
            f"R{row.value:,.2f} the dockets show, R{abs(shortfall):,.2f} out: "
            f"R{row.value:,.2f} does not divide by {row.cartons.net:g} cartons at any price."
        )
    return rounded


def _last_written_row(sheet: Any, layout: SheetLayout) -> int:
    """The last row with anything in it, ignoring the blank tail Excel leaves behind."""
    last = layout.header_row
    for row_number in range(sheet.max_row, layout.header_row, -1):
        if any(sheet.cell(row_number, c).value is not None for c in range(1, sheet.max_column + 1)):
            last = row_number
            break
    return last


def _formula(field_name: str, layout: SheetLayout, row_number: int) -> str | None:
    """Build one formula from the letters this book actually uses.

    Returns `None` when a column the formula points at is missing, because a formula referring to
    a column that is not there is worse than an empty cell: it opens as `#REF!` and looks like
    corruption in a file the business settles money against.
    """
    letters = {name: letter for name in layout.columns if (letter := layout.letter(name))}
    try:
        return FORMULAS[field_name].format(r=row_number, **letters)
    except KeyError:
        return None


def append(path: Path, built: AppendPlan, label: str = "") -> AppendResult:
    """Snapshot, write, and either keep both or neither (D4).

    The copy is taken first and put back if anything goes wrong, so a half-written book is not a
    state this can leave behind. The caller holds the database transaction open across this call
    and rolls it back on the exception, which is what keeps "appended" and "actually in the file"
    the same answer.
    """
    if not built.is_writable:
        raise AppendRefusedError("; ".join(built.refusals) or "There is nothing to append.")
    if not path.exists():
        raise AppendRefusedError(f"There is no workbook at {path}. Nothing was written.")

    layout = locate(path)
    saved = snapshot.take(path, label=label or "before append")
    temporary = path.with_suffix(".appending.xlsx")
    try:
        workbook = load_workbook(path, data_only=False)
        try:
            sheet = workbook[layout.sheet_name]
            template_row = _last_written_row(sheet, layout)
            start = template_row + 1

            for offset, planned in enumerate(built.rows):
                row_number = start + offset
                for name, index in layout.columns.items():
                    if name in NEVER_WRITTEN:
                        continue
                    cell = sheet.cell(row_number, index)
                    if name in FORMULAS:
                        formula = _formula(name, layout, row_number)
                        if formula is not None:
                            cell.value = formula
                    else:
                        cell.value = _cell_value(planned.values.get(name))
                    if template_row > layout.header_row:
                        # Number formats must match the existing sheet: the price column carries
                        # five decimals and the date column `d-mmm`, and a row that renders
                        # differently from its neighbours reads as an error in the figures.
                        cell.number_format = sheet.cell(template_row, index).number_format

            workbook.save(temporary)
        finally:
            workbook.close()
        os.replace(temporary, path)
    except Exception:
        Path(temporary).unlink(missing_ok=True)
        if saved.path.exists():
            os.replace(saved.path, path)
        raise
    snapshot.prune()
    return AppendResult(
        first_row=start,
        last_row=start + len(built.rows) - 1,
        saved_as=saved,
        sheet_name=layout.sheet_name,
        letters={name: layout.letter(name) or "" for name in layout.columns},
    )


def _cell_value(value: object) -> object:
    """openpyxl writes `Decimal` as a number of its own accord, so nothing here becomes a float."""
    if value is None or isinstance(value, (Decimal, int, str, date)):
        return value
    return str(value)


__all__ = [
    "FORMULAS",
    "AppendPlan",
    "AppendRefusedError",
    "AppendResult",
    "PlannedRow",
    "WorkbookShapeError",
    "append",
    "plan",
]
