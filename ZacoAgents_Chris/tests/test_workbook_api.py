"""Ingest to appended, end to end over HTTP (section 5, D4).

`tests/test_workbook_append.py` pins the cells. This pins the path an operator actually takes:
the queue closes, the preview shows what would happen, the append happens once, and the book
before it is recoverable.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from tests.test_accounts_api import PASSWORD, _make_user
from zaco.auth.permissions import Permission
from zaco.resolve.service import workbook_path

pytestmark = pytest.mark.db

DATA = Path(__file__).resolve().parent.parent / "data"
SALES = "DailySalesDetail_20260525-20260531.csv"
PAYMENTS = "PaymentDetails_20260529-20260602.csv"


@pytest.fixture(autouse=True)
def book_is_put_back() -> Any:
    """Every case here writes to the shared test workbook, so every case hands it back."""
    path = workbook_path()
    before = path.read_bytes()
    yield path
    path.write_bytes(before)


@pytest.fixture
def operator(client: TestClient, db: Session) -> TestClient:
    _make_user(
        db,
        "operator@example.com",
        [Permission.INGEST, Permission.RESOLVE, Permission.APPEND],
    )
    client.post("/api/auth/login", json={"email": "operator@example.com", "password": PASSWORD})
    return client


def _settle(client: TestClient, *names: str) -> int:
    """Upload a round, answer every question, and close the queue."""
    files = [("files", (n, (DATA / n).read_bytes(), "text/csv")) for n in names]
    response = client.post("/api/rounds", files=files)
    assert response.status_code == 201, response.text
    round_id = response.json()["summary"]["id"]

    while True:
        body = client.get(f"/api/rounds/{round_id}").json()
        if not body["queue"]:
            break
        item = body["queue"][0]
        if item["kind"] == "product_code":
            client.post(
                "/api/products/code",
                json={"product_key": item["key"], "short_code": "Imp " + item["title"][:14]},
            )
        elif item["kind"] == "product_link":
            left, right = item["key"].split("||")
            client.post(
                "/api/products/link",
                json={"left": left, "right": right, "accepted": False, "reason": "not proven"},
            )
        else:
            client.post(
                f"/api/rounds/{round_id}/delivery-notes",
                params={"delivery_id": item["key"]},
                json={"dn": item["proposal"], "provenance": "operator", "reason": "as offered"},
            )
    assert client.post(f"/api/rounds/{round_id}/resolve").status_code == 200
    return round_id


# --- what the book says about itself ---------------------------------------------------------------


def test_the_state_reports_the_letters_this_book_actually_uses(operator: TestClient) -> None:
    """Not decoration. The brief's letters are wrong for the real file, and the only way an
    operator can check the system is writing where they think is to see what it resolved."""
    body = operator.get("/api/workbook").json()
    assert body["is_readable"] is True
    assert body["sheet_name"] == "Sheet1"
    assert body["letters"]["baby_stock"] == "L"
    assert body["letters"]["dn"] == "A"
    assert body["unknown_headers"] == {"Buyer note": 3, "Packhouse": 22}


def test_a_round_only_reaches_the_ready_list_once_its_queue_is_closed(
    operator: TestClient,
) -> None:
    files = [("files", (SALES, (DATA / SALES).read_bytes(), "text/csv"))]
    round_id = operator.post("/api/rounds", files=files).json()["summary"]["id"]
    assert operator.get("/api/workbook").json()["ready_rounds"] == []

    body = operator.get(f"/api/rounds/{round_id}/append").json()
    assert body["is_writable"] is False
    assert any("queue" in r for r in body["refusals"])


# --- the preview -----------------------------------------------------------------------------------


def test_the_preview_shows_the_formulas_and_writes_nothing(operator: TestClient) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    before = workbook_path().read_bytes()

    body = operator.get(f"/api/rounds/{round_id}/append").json()
    assert body["is_writable"] is True
    assert body["first_row"] == 5
    assert body["rows"], "the preview showed no rows"
    assert body["rows"][0]["cells"]["baby_stock"] == "=I5-K5"
    assert body["rows"][0]["cells"]["gross_total"] == "=SUM(K5*M5)"
    assert body["rows"][0]["cells"]["completed"] == "Incomplete"
    assert body["rows"][0]["why"] is not None
    assert workbook_path().read_bytes() == before, "the preview touched the file"


def test_the_preview_names_the_columns_it_will_not_write(operator: TestClient) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    body = operator.get(f"/api/rounds/{round_id}/append").json()
    assert body["never_written"] == ["notes"]
    assert set(body["formula_columns"]) == {
        "baby_stock",
        "frui_curr_sales_value",
        "frui_price_per_crt",
        "gross_total",
        "markup_percent",
        "nett_price_per_crt",
        "z_rand_per_crt",
        "z_total",
    }
    assert all(row["cells"]["notes"] == "" for row in body["rows"])


def test_qty_received_is_written_once_per_consignment_and_never_guessed(
    operator: TestClient,
) -> None:
    """What was sent is sent once, however many account sales it sells under -- and where no
    document says what was sent, the cell is blank with a reason rather than nought."""
    round_id = _settle(operator, SALES, PAYMENTS)
    rows = operator.get(f"/api/rounds/{round_id}/append").json()["rows"]

    for row in rows:
        if not row["cells"]["qty_received"]:
            reason = row["blanks"].get("qty_received", "")
            assert reason, f"{row['delivery_id']} has a blank Qty Received and no reason given"
            assert reason.startswith("counted on row") or reason == "not reported", reason
    assert not any(row["cells"]["qty_received"] == "0" for row in rows), "absent was written as 0"


def test_what_was_sent_is_read_off_the_consignment_report_when_there_is_one(
    operator: TestClient,
) -> None:
    """The daily sales file cannot say what was sent; the consignment report can. With both in
    the round, Qty Received is a figure rather than a blank."""
    round_id = _settle(operator, SALES, "ConsignmentReports_20260525-20260531.txt", PAYMENTS)
    rows = operator.get(f"/api/rounds/{round_id}/append").json()["rows"]
    assert any(row["cells"]["qty_received"] for row in rows), "nothing carried what was sent"


# --- what the grid needs -----------------------------------------------------------------------


def test_the_preview_carries_the_book_s_own_headers_in_the_book_s_own_order(
    operator: TestClient,
) -> None:
    """The grid shows the operator their own words above their own letters. Prettifying the
    system's field names instead would hide a header that does not say what they think."""
    round_id = _settle(operator, SALES, PAYMENTS)
    body = operator.get(f"/api/rounds/{round_id}/append").json()

    assert body["headers"]["baby_stock"] == "Baby Stock"
    assert body["headers"]["notes"] == "NOTES"
    assert body["order"][:3] == ["dn", "market_agent", "completed"]
    assert [body["letters"][f] for f in body["order"][:3]] == ["A", "B", "D"]
    assert "price" in body["numeric_columns"]
    assert "market_agent" not in body["numeric_columns"]


def test_a_row_that_cannot_be_written_is_still_shown_with_what_is_missing(
    operator: TestClient,
) -> None:
    """Dropping it and listing the reason separately gives a short grid and a long complaint, and
    leaves the operator working out which row the complaint belongs to."""
    files = [("files", (SALES, (DATA / SALES).read_bytes(), "text/csv"))]
    round_id = operator.post("/api/rounds", files=files).json()["summary"]["id"]

    body = operator.get(f"/api/rounds/{round_id}/append").json()
    assert body["is_writable"] is False
    assert body["rows"], "the preview showed nothing at all"

    stuck = [r for r in body["rows"] if not r["is_writable"]]
    assert stuck, "every row claimed to be writable on an unanswered round"
    assert all(r["blocked_by"] for r in stuck)
    assert stuck[0]["blanks"]["dn"] == "DN not captured"
    assert stuck[0]["blanks"]["description"] == "code unmapped"
    assert stuck[0]["cells"]["dn"] == ""


def test_every_blank_cell_carries_its_own_reason(operator: TestClient) -> None:
    """A blank nobody explained is indistinguishable from one nobody got to."""
    round_id = _settle(operator, SALES, PAYMENTS)
    body = operator.get(f"/api/rounds/{round_id}/append").json()

    unexplained = [
        (row["row_number"], name)
        for row in body["rows"]
        for name in body["order"]
        if name not in body["never_written"]
        and name not in body["formula_columns"]
        and not row["cells"][name]
        and not row["blanks"].get(name)
    ]
    assert not unexplained, f"blank with no reason: {unexplained}"


def test_a_nett_split_across_rows_says_so_in_the_cell(operator: TestClient) -> None:
    """Section 8 apportions it. Until then the cell says why, rather than showing nothing."""
    round_id = _settle(operator, SALES, PAYMENTS)
    labels = {
        row["blanks"].get("nett_total", "")
        for row in operator.get(f"/api/rounds/{round_id}/append").json()["rows"]
    }
    assert any("split across" in label or "no payment run" in label for label in labels), labels


# --- the append ------------------------------------------------------------------------------------


def test_appending_writes_the_rows_and_leaves_everything_else_alone(
    operator: TestClient,
) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    kept = {
        cell.coordinate: cell.value
        for row in load_workbook(workbook_path())["Sheet1"].iter_rows(min_row=1, max_row=4)
        for cell in row
    }

    response = operator.post(f"/api/rounds/{round_id}/append")
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "appended"
    assert body["appended_rows"] == f"5-{4 + len(body['rows'])}"
    assert body["saved_as"], "no version was kept"

    sheet = load_workbook(workbook_path())["Sheet1"]
    for coordinate, value in kept.items():
        assert sheet[coordinate].value == value, f"{coordinate} changed"
    assert sheet["D5"].value == "Incomplete"
    assert sheet["L5"].value == "=I5-K5"
    assert sheet["W5"].value is None


def test_a_round_cannot_be_appended_twice(operator: TestClient) -> None:
    """Rows are appended, never rebuilt, so a second append would write every row again."""
    round_id = _settle(operator, SALES, PAYMENTS)
    assert operator.post(f"/api/rounds/{round_id}/append").status_code == 200
    rows = load_workbook(workbook_path())["Sheet1"].max_row

    again = operator.post(f"/api/rounds/{round_id}/append")
    assert again.status_code == 409
    assert "already appended" in again.json()["detail"]
    assert load_workbook(workbook_path())["Sheet1"].max_row == rows


def test_an_appended_round_cannot_be_reopened(operator: TestClient) -> None:
    """An append cannot be unwritten, so the book is corrected by rolling it back, not the round."""
    round_id = _settle(operator, SALES, PAYMENTS)
    operator.post(f"/api/rounds/{round_id}/append")

    refused = operator.post(f"/api/rounds/{round_id}/reopen", json={"reason": "wrong file"})
    assert refused.status_code == 409
    assert "rolling it back" in refused.json()["detail"]


def test_the_append_is_recorded_against_the_round(operator: TestClient) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    operator.post(f"/api/rounds/{round_id}/append")

    events = operator.get(f"/api/rounds/{round_id}").json()["events"]
    written = next(e for e in events if e["action"] == "workbook_appended")
    assert "Sheet1" in written["subject"]
    assert written["by"] == "operator@example.com"

    state = operator.get("/api/workbook").json()
    assert [r["round_id"] for r in state["appended_rounds"]] == [round_id]
    assert state["ready_rounds"] == []


# --- the version kept, and the rollback ------------------------------------------------------------


def test_the_version_before_the_append_is_kept_and_can_be_put_back(
    operator: TestClient,
) -> None:
    before = workbook_path().read_bytes()
    round_id = _settle(operator, SALES, PAYMENTS)
    saved = operator.post(f"/api/rounds/{round_id}/append").json()["saved_as"]
    assert workbook_path().read_bytes() != before

    restored = operator.post(
        f"/api/workbook/versions/{saved}/restore",
        json={"reason": "appended against the wrong producer"},
    )
    assert restored.status_code == 200, restored.text
    assert workbook_path().read_bytes() == before


def test_rolling_back_needs_a_reason(operator: TestClient) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    saved = operator.post(f"/api/rounds/{round_id}/append").json()["saved_as"]
    response = operator.post(f"/api/workbook/versions/{saved}/restore", json={"reason": "  "})
    assert response.status_code == 422


def test_a_version_that_is_not_one_of_ours_is_not_found(operator: TestClient) -> None:
    response = operator.post(
        "/api/workbook/versions/..%2F..%2Faccount-sales-book.xlsx/restore",
        json={"reason": "trying it on"},
    )
    assert response.status_code == 404


def test_a_rollback_does_not_quietly_reopen_the_round_it_undid(operator: TestClient) -> None:
    """The rows may or may not be in the version restored, and reopening would invite a second
    append that duplicates whatever survived."""
    round_id = _settle(operator, SALES, PAYMENTS)
    saved = operator.post(f"/api/rounds/{round_id}/append").json()["saved_as"]
    operator.post(f"/api/workbook/versions/{saved}/restore", json={"reason": "wrong producer"})

    assert operator.get(f"/api/rounds/{round_id}").json()["summary"]["status"] == "appended"
    assert operator.post(f"/api/rounds/{round_id}/append").status_code == 409


def test_a_rollback_that_removed_the_rows_is_reported_against_the_round(
    operator: TestClient,
) -> None:
    """Keeping the appended mark is right and it leaves the operator blind. So say it."""
    round_id = _settle(operator, SALES, PAYMENTS)
    saved = operator.post(f"/api/rounds/{round_id}/append").json()["saved_as"]

    before = operator.get("/api/workbook").json()
    mine = next(r for r in before["appended_rounds"] if r["round_id"] == round_id)
    assert mine["agrees"] is True
    assert mine["finding"] is None

    operator.post(f"/api/workbook/versions/{saved}/restore", json={"reason": "wrong producer"})

    after = operator.get("/api/workbook").json()
    mine = next(r for r in after["appended_rounds"] if r["round_id"] == round_id)
    assert mine["agrees"] is False
    assert mine["finding"] is not None
    assert f"rows {mine['first_row']}-{mine['last_row']}" in mine["finding"]
    assert "no rows there at all" in mine["finding"]


def test_the_agreement_states_what_it_did_not_compare(operator: TestClient) -> None:
    """Section 10: the blind spot travels with the figure, not in a comment."""
    round_id = _settle(operator, SALES, PAYMENTS)
    operator.post(f"/api/rounds/{round_id}/append")

    mine = next(
        r
        for r in operator.get("/api/workbook").json()["appended_rounds"]
        if r["round_id"] == round_id
    )
    assert "figures in those rows are not compared" in mine["checked"]


def test_an_unappended_round_is_not_held_against_the_book_at_all(operator: TestClient) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)

    state = operator.get("/api/workbook").json()

    assert [r["round_id"] for r in state["ready_rounds"]] == [round_id]
    assert state["appended_rounds"] == []


def test_an_appended_round_previews_at_the_rows_it_wrote_not_the_next_free_one(
    operator: TestClient,
) -> None:
    """The row number goes into every formula, so the wrong one is not a cosmetic slip."""
    round_id = _settle(operator, SALES, PAYMENTS)
    written = operator.post(f"/api/rounds/{round_id}/append").json()
    first, last = written["first_row"], int(written["appended_rows"].split("-")[1])

    again = operator.get(f"/api/rounds/{round_id}/append").json()

    assert again["first_row"] == first
    assert [r["row_number"] for r in again["rows"]] == [str(n) for n in range(first, last + 1)]


def test_the_formulas_of_an_appended_round_match_the_cells_the_book_holds(
    operator: TestClient, book_is_put_back: Path
) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    operator.post(f"/api/rounds/{round_id}/append")

    preview = operator.get(f"/api/rounds/{round_id}/append").json()
    sheet = load_workbook(book_is_put_back)[preview["headers"] and "Sheet1"]
    for row in preview["rows"]:
        for field in preview["formula_columns"]:
            letter = preview["letters"][field]
            assert sheet[f"{letter}{row['row_number']}"].value == row["cells"][field]


def test_an_appended_round_names_the_copy_taken_before_it(operator: TestClient) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    written = operator.post(f"/api/rounds/{round_id}/append").json()

    again = operator.get(f"/api/rounds/{round_id}/append").json()

    assert again["saved_as"] == written["saved_as"]
    assert again["saved_as"]


# --- who may do it -----------------------------------------------------------------------------------


def test_closing_a_queue_does_not_carry_the_right_to_write_the_book(
    client: TestClient, db: Session, operator: TestClient
) -> None:
    round_id = _settle(operator, SALES, PAYMENTS)
    _make_user(db, "resolver@example.com", [Permission.INGEST, Permission.RESOLVE])
    client.post("/api/auth/login", json={"email": "resolver@example.com", "password": PASSWORD})

    assert client.post(f"/api/rounds/{round_id}/append").status_code == 403
    assert client.get(f"/api/rounds/{round_id}/append").status_code == 403
    assert client.get("/api/workbook").status_code == 200, "reading the state is not writing it"


def test_the_stm_column_follows_d7_even_with_no_payment_run_behind_the_row(
    operator: TestClient,
) -> None:
    """D7 belongs to the column, not to the payment run. A row whose account sale appears in no
    payment document still goes into the operator's STM No column and still has to look like the
    rows around it -- `PRE*BT*390100` sitting between two bare numbers stops a filter working.

    `JOH*SUB*5644200/1` is the other half of the same rule: it has no bare number, so it keeps
    the whole reference. Dropping to `5644200/1` would lose which agent ran it, and `5640001/1`
    and `5640001/2` are already two separate April runs worth R5,100 and R3,230.
    """
    round_id = _settle(operator, SALES)
    values = [
        row["cells"]["stm_no"]
        for row in operator.get(f"/api/rounds/{round_id}/append").json()["rows"]
    ]
    assert values, "the fixture produced no rows"
    for value in values:
        bare = value.rsplit("*", 1)[-1]
        assert not bare.isdigit() or "*" not in value, (
            f"{value} kept its reference prefix even though it has a bare number"
        )
