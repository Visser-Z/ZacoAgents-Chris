"""Appending to the operator's live book (section 5).

The golden file here is the operator's **real** 23-column book, copied per test. Asserting cell
by cell rather than against a committed expected `.xlsx` is deliberate: a binary comparison says
"different" and a cell comparison says which cell, and the whole risk in this phase is writing
the right value into the wrong column.

The letters matter. The brief prints `Baby Stock` as `=H{r}-J{r}`; in the real book, where
`Buyer note` sits at C, it is `=I{r}-K{r}`. A test that pinned the brief's letters would pass on
a file no operator has.
"""

from __future__ import annotations

import shutil
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from zaco.workbook import snapshot
from zaco.workbook.append import (
    FORMULAS,
    AppendPlan,
    AppendRefusedError,
    PlannedRow,
    append,
)
from zaco.workbook.locate import WorkbookShapeError, locate

REPO_ROOT = Path(__file__).resolve().parent.parent
LIVE = REPO_ROOT / "workbook" / "account-sales-book.xlsx"

#: The letters the real book actually uses. Not the brief's.
REAL = {
    "dn": "A",
    "market_agent": "B",
    "completed": "D",
    "date": "E",
    "stm_no": "F",
    "description": "G",
    "qty_received": "H",
    "opening_stock": "I",
    "frui_price_per_crt": "J",
    "cartons_sold": "K",
    "baby_stock": "L",
    "price": "M",
    "gross_total": "N",
    "nett_total": "O",
    "nett_price_per_crt": "P",
    "z_rand_per_crt": "Q",
    "z_total": "R",
    "markup_percent": "S",
    "frui_curr_sales_value": "T",
    "status": "U",
    "notes": "W",
}


@pytest.fixture
def book(tmp_path: Path) -> Path:
    """A copy of the operator's real book. Never the original."""
    target = tmp_path / "account-sales-book.xlsx"
    shutil.copy(LIVE, target)
    return target


def _row(**values: object) -> PlannedRow:
    planned = PlannedRow(
        delivery_id=str(values.pop("delivery_id", "2200100Z")),
        consignment_id="C1",
        product="LEMONS EUREKA",
        account_sale=str(values.pop("account_sale", "PRE*BT*390100")),
    )
    planned.values = {
        "dn": 14930,
        "market_agent": "Farmers Trust",
        "completed": "Incomplete",
        "date": date(2026, 7, 6),
        "stm_no": 390100,
        "description": "Imp Lemons 15kg",
        "qty_received": Decimal("80"),
        "opening_stock": Decimal("80"),
        "cartons_sold": Decimal("30"),
        "price": Decimal("30.00"),
        "nett_total": Decimal("765.00"),
        "status": "10.07",
        **values,
    }
    return planned


def _plan(*rows: PlannedRow) -> AppendPlan:
    return AppendPlan(rows=list(rows) or [_row()])


# --- what the file keeps -------------------------------------------------------------------------


def test_the_other_sheets_are_left_exactly_as_they_were(book: Path) -> None:
    before = load_workbook(book)
    kept = {
        name: [list(r) for r in before[name].iter_rows(values_only=True)]
        for name in ("Cover", "Rates")
    }
    before.close()

    append(book, _plan())

    after = load_workbook(book)
    assert after.sheetnames == ["Cover", "Rates", "Sheet1"]
    for name, rows in kept.items():
        assert [list(r) for r in after[name].iter_rows(values_only=True)] == rows


def test_the_rows_that_were_there_are_untouched(book: Path) -> None:
    """Rows are appended, never rebuilt, and that has to be true of the formats too."""
    before = load_workbook(book)["Sheet1"]
    kept = {
        cell.coordinate: (cell.value, cell.number_format)
        for row in before.iter_rows(min_row=1, max_row=4)
        for cell in row
    }

    append(book, _plan())

    after = load_workbook(book)["Sheet1"]
    for coordinate, (value, number_format) in kept.items():
        cell = after[coordinate]
        assert cell.value == value, f"{coordinate} changed"
        assert cell.number_format == number_format, f"{coordinate} lost its format"


def test_the_two_columns_nothing_here_owns_are_not_written(book: Path) -> None:
    """`Buyer note` at C and `Packhouse` at V belong to the operator's own use of the file."""
    append(book, _plan())
    sheet = load_workbook(book)["Sheet1"]
    assert sheet["C5"].value is None
    assert sheet["V5"].value is None


def test_the_notes_column_is_never_written(book: Path) -> None:
    """Section 5: read back and written back unchanged, which means left alone."""
    append(book, _plan())
    sheet = load_workbook(book)["Sheet1"]
    assert sheet["W2"].value == "FT Received 18"
    assert sheet["W4"].value == "FT HET NIE INGENEEM NIE"
    assert sheet["W5"].value is None


# --- the formulas --------------------------------------------------------------------------------


def test_the_formulas_use_this_book_s_letters_and_not_the_brief_s(book: Path) -> None:
    """The single most expensive mistake available in this phase.

    The brief says `Baby Stock` is `=H{r}-J{r}`. Written into the real book that would subtract
    `Frui Price/crt` from `Qty Received` and put the answer where the stock left over goes -- a
    file that opens without complaint and is wrong in every row.
    """
    append(book, _plan())
    sheet = load_workbook(book)["Sheet1"]

    assert sheet["J5"].value == '=IFERROR(P5*70%,"-")'
    assert sheet["L5"].value == "=I5-K5"
    assert sheet["N5"].value == "=SUM(K5*M5)"
    assert sheet["P5"].value == '=IFERROR(O5/K5,"-")'
    assert sheet["Q5"].value == '=IFERROR(P5-J5,"-")'
    assert sheet["R5"].value == "=O5-T5"
    assert sheet["S5"].value == '=IFERROR(Q5/P5,"-")'
    assert sheet["T5"].value == "=J5*K5"


def test_each_appended_row_gets_its_own_row_number(book: Path) -> None:
    append(book, _plan(_row(), _row(account_sale="PRE*BT*390110"), _row()))
    sheet = load_workbook(book)["Sheet1"]
    assert [sheet[f"L{r}"].value for r in (5, 6, 7)] == ["=I5-K5", "=I6-K6", "=I7-K7"]


def test_the_formula_columns_are_the_ones_the_brief_names(book: Path) -> None:
    assert set(FORMULAS) == {
        "frui_price_per_crt",
        "baby_stock",
        "gross_total",
        "nett_price_per_crt",
        "z_rand_per_crt",
        "z_total",
        "markup_percent",
        "frui_curr_sales_value",
    }


def test_no_computed_value_reaches_a_formula_column(book: Path) -> None:
    """A row where every formula's inputs are known still gets formulas, not answers."""
    append(book, _plan())
    sheet = load_workbook(book)["Sheet1"]
    for field_name in FORMULAS:
        cell = sheet[f"{REAL[field_name]}5"]
        assert isinstance(cell.value, str) and cell.value.startswith("="), (
            f"{field_name} was written as a value"
        )


# --- the written columns -------------------------------------------------------------------------


def test_the_written_cells_land_in_the_columns_the_headers_name(book: Path) -> None:
    append(book, _plan())
    sheet = load_workbook(book)["Sheet1"]
    assert sheet["A5"].value == 14930
    assert sheet["B5"].value == "Farmers Trust"
    assert sheet["D5"].value == "Incomplete"
    assert sheet["E5"].value.date() == date(2026, 7, 6)
    assert sheet["F5"].value == 390100
    assert sheet["G5"].value == "Imp Lemons 15kg"
    assert sheet["H5"].value == 80
    assert sheet["I5"].value == 80
    assert sheet["K5"].value == 30
    assert sheet["M5"].value == Decimal("30.00")
    assert sheet["O5"].value == Decimal("765.00")
    assert sheet["U5"].value == "10.07"


def test_money_reaches_the_file_to_the_cent(book: Path) -> None:
    """Every currency column is two decimals, and lands as exactly those two decimals.

    Read back as a float, because that is what xlsx stores and what Excel will show. The
    assertion is that the file holds the cent, not that openpyxl hands back a `Decimal`.
    """
    append(book, _plan(_row(nett_total=Decimal("1234.57"), price=Decimal("41.15"))))
    sheet = load_workbook(book)["Sheet1"]
    assert str(sheet["O5"].value) == "1234.57"
    assert str(sheet["M5"].value) == "41.15"


def test_a_price_is_written_at_a_precision_the_file_can_actually_hold(book: Path) -> None:
    """`openpyxl` turns a `Decimal` into a float on the way in, so a 28-digit price becomes
    `133.33333333333331` whatever this system does. Rounding to the five decimals the column
    already displays writes a number that survives the trip, rather than one that changes
    quietly between here and the file."""
    from zaco.workbook.append import PRICE_PLACES

    assert Decimal("0.00001") == PRICE_PLACES
    append(book, _plan(_row(price=Decimal("133.33333"))))
    assert str(load_workbook(book)["Sheet1"]["M5"].value) == "133.33333"


def test_number_formats_match_the_rows_above(book: Path) -> None:
    """A row that renders differently from its neighbours reads as an error in the figures."""
    append(book, _plan())
    sheet = load_workbook(book)["Sheet1"]
    assert sheet["E5"].number_format == sheet["E4"].number_format == "d-mmm"
    assert sheet["M5"].number_format == "#,##0.00000"
    assert sheet["O5"].number_format == "#,##0.00"


def test_a_recorded_no_delivery_note_writes_a_blank_column_a(book: Path) -> None:
    """D11: visibly empty, with the reason held in the system, is a different thing from unfinished."""
    append(book, _plan(_row(dn=None)))
    sheet = load_workbook(book)["Sheet1"]
    assert sheet["A5"].value is None
    assert sheet["B5"].value == "Farmers Trust"


def test_an_account_sale_with_no_bare_number_keeps_its_suffix(book: Path) -> None:
    """`5640001/1` and `5640001/2` are two April payment runs worth R5,100 and R3,230."""
    append(book, _plan(_row(stm_no="5640001/1")))
    assert load_workbook(book)["Sheet1"]["F5"].value == "5640001/1"


# --- where it appends ----------------------------------------------------------------------------


def test_rows_go_beneath_what_is_already_there(book: Path) -> None:
    result = append(book, _plan(_row(), _row()))
    assert (result.first_row, result.last_row) == (5, 6)
    assert load_workbook(book)["Sheet1"].max_row == 6


def test_a_second_append_goes_beneath_the_first(book: Path) -> None:
    append(book, _plan())
    result = append(book, _plan())
    assert result.first_row == 6
    sheet = load_workbook(book)["Sheet1"]
    assert sheet["L5"].value == "=I5-K5"
    assert sheet["L6"].value == "=I6-K6"


def test_the_sheet_is_found_by_its_headers_even_though_it_is_third(book: Path) -> None:
    result = append(book, _plan())
    assert result.sheet_name == "Sheet1"
    assert result.letters["baby_stock"] == "L"


def test_a_reordered_book_is_written_correctly_because_nothing_counts_positions(
    tmp_path: Path,
) -> None:
    """Real workbooks do not match the template. The letters follow the headers, wherever they go."""
    source = load_workbook(LIVE)["Sheet1"]
    headers = [c.value for c in source[1]]
    reordered = list(reversed(headers))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Whatever"
    sheet.append(reordered)
    sheet.append([None] * len(reordered))
    path = tmp_path / "reordered.xlsx"
    workbook.save(path)

    result = append(path, _plan())
    layout = locate(path)
    written = load_workbook(path)[layout.sheet_name]
    baby = layout.letter("baby_stock")
    opening = layout.letter("opening_stock")
    sold = layout.letter("cartons_sold")
    assert written[f"{baby}{result.first_row}"].value == (
        f"={opening}{result.first_row}-{sold}{result.first_row}"
    )
    assert written[f"{layout.letter('description')}{result.first_row}"].value == "Imp Lemons 15kg"


def test_a_file_with_no_recognisable_data_sheet_is_refused_with_an_explanation(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    workbook.active.append(["something", "else", "entirely"])
    path = tmp_path / "not-the-book.xlsx"
    workbook.save(path)

    with pytest.raises(WorkbookShapeError) as refused:
        append(path, _plan())
    assert "no sheet with a recognisable header row" in str(refused.value)
    assert "Nothing was written" in str(refused.value)


def test_a_missing_workbook_is_refused_before_anything_is_opened(tmp_path: Path) -> None:
    with pytest.raises(AppendRefusedError) as refused:
        append(tmp_path / "nowhere.xlsx", _plan())
    assert "Nothing was written" in str(refused.value)


def test_a_plan_that_refuses_is_never_written(book: Path) -> None:
    before = book.read_bytes()
    built = _plan()
    built.refusals.append("2200100Z has no approved delivery note")
    with pytest.raises(AppendRefusedError) as refused:
        append(book, built)
    assert "no approved delivery note" in str(refused.value)
    assert book.read_bytes() == before


# --- the snapshot and the rollback (D4) ------------------------------------------------------------


def test_the_book_is_copied_aside_in_the_same_step_as_the_append(book: Path) -> None:
    before = book.read_bytes()
    result = append(book, _plan())
    assert result.saved_as.path.exists()
    assert result.saved_as.path.read_bytes() == before
    assert book.read_bytes() != before


def test_rolling_back_restores_the_earlier_version_byte_for_byte(book: Path) -> None:
    before = book.read_bytes()
    result = append(book, _plan())
    assert book.read_bytes() != before

    snapshot.restore(result.saved_as, book)
    assert book.read_bytes() == before


def test_rolling_back_keeps_the_version_it_replaced(book: Path) -> None:
    """A one-click control that can destroy the file somebody was working on is a trap."""
    first = append(book, _plan())
    appended = book.read_bytes()

    replaced = snapshot.restore(first.saved_as, book)
    assert replaced.path.read_bytes() == appended
    assert "rollback" in replaced.label


def test_a_write_that_fails_puts_the_book_back(book: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Snapshot, append, and either both or neither -- there is no half-written state to land in."""
    before = book.read_bytes()

    def explode(self: object, *args: object, **kwargs: object) -> None:
        raise OSError("the disk went away")

    monkeypatch.setattr("openpyxl.workbook.workbook.Workbook.save", explode)
    with pytest.raises(OSError):
        append(book, _plan())
    assert book.read_bytes() == before


def test_snapshots_are_listed_newest_first_and_pruned_to_the_retention(book: Path) -> None:
    for _ in range(3):
        snapshot.take(book, label="a check")
    everything = snapshot.listing()
    assert everything == sorted(everything, key=lambda s: s.taken_at, reverse=True)

    snapshot.prune(keep=1)
    assert len(snapshot.listing()) == 1


def test_a_snapshot_is_only_found_by_a_name_that_is_one_of_ours(book: Path) -> None:
    taken = snapshot.take(book, label="a check")
    assert snapshot.find(taken.name) is not None
    assert snapshot.find("../../etc/passwd") is None
    assert snapshot.find("account-sales-book.xlsx") is None
